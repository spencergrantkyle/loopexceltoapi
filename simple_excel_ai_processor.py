#!/usr/bin/env python3
"""
Simple Excel AI Processor
Processes Excel cell ranges through OpenAI API and outputs results to a new Excel file.

Usage:
    python simple_excel_ai_processor.py
"""

import os
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openai import OpenAI
from typing import List, Dict, Tuple
import time

# Load environment variables from .env file if it exists
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass  # dotenv is optional

class SimpleExcelAIProcessor:
    def __init__(self, api_key: str = None):
        """
        Initialize the processor with OpenAI API key
        
        Args:
            api_key: OpenAI API key (if None, will try to get from environment)
        """
        if api_key:
            self.client = OpenAI(api_key=api_key)
        else:
            self.client = OpenAI()  # Will use OPENAI_API_KEY environment variable
        
        self.results = []
    
    def get_cell_value_and_formula(self, workbook_path: str, sheet_name: str, cell_ref: str) -> Tuple[str, str]:
        """
        Get both the value and formula of a cell
        
        Args:
            workbook_path: Path to the Excel file
            sheet_name: Name of the sheet
            cell_ref: Cell reference (e.g., 'A1', 'B2')
            
        Returns:
            Tuple of (cell_value, cell_formula)
        """
        try:
            wb = load_workbook(workbook_path, data_only=False)  # data_only=False to get formulas
            ws = wb[sheet_name]
            
            # Get the cell
            cell = ws[cell_ref]
            
            # Get formula if it exists, otherwise get value
            cell_formula = cell.value if cell.data_type == 'f' else None
            cell_value = cell.value
            
            # If it's a formula, get the calculated value
            if cell_formula:
                wb_data = load_workbook(workbook_path, data_only=True)  # data_only=True to get calculated values
                ws_data = wb_data[sheet_name]
                cell_data = ws_data[cell_ref]
                cell_value = cell_data.value
            
            return str(cell_value) if cell_value is not None else "", str(cell_formula) if cell_formula else ""
            
        except Exception as e:
            print(f"Error reading cell {cell_ref} in sheet {sheet_name}: {e}")
            return "", ""
    
    def process_cell_range(self, workbook_path: str, sheet_name: str, cell_range: str, 
                          custom_prompt: str, model: str = "gpt-4o") -> List[Dict]:
        """
        Process a range of cells through OpenAI API
        
        Args:
            workbook_path: Path to the Excel file
            sheet_name: Name of the sheet
            cell_range: Cell range (e.g., 'A1:B10')
            custom_prompt: Custom prompt to send with each cell value
            model: OpenAI model to use
            
        Returns:
            List of results with cell info and API responses
        """
        try:
            wb = load_workbook(workbook_path)
            ws = wb[sheet_name]
            
            # Parse the cell range
            if ':' in cell_range:
                start_cell, end_cell = cell_range.split(':')
                cells = list(ws[start_cell:end_cell])
            else:
                # Single cell
                cells = [[ws[cell_range]]]
            
            results = []
            
            for row in cells:
                for cell in row:
                    cell_ref = cell.coordinate
                    cell_value, cell_formula = self.get_cell_value_and_formula(workbook_path, sheet_name, cell_ref)
                    
                    if cell_value.strip():  # Only process non-empty cells
                        print(f"Processing cell {cell_ref}...")
                        
                        # Call OpenAI API
                        try:
                            response = self.client.chat.completions.create(
                                model=model,
                                messages=[
                                    {
                                        "role": "system",
                                        "content": "You are a helpful assistant that processes Excel cell content according to the user's instructions."
                                    },
                                    {
                                        "role": "user",
                                        "content": f"{custom_prompt}\n\nCell Value: {cell_value}\nCell Formula: {cell_formula if cell_formula else 'No formula'}"
                                    }
                                ],
                                max_tokens=1000,
                                temperature=0.3
                            )
                            
                            api_response = response.choices[0].message.content.strip()
                            
                            results.append({
                                'sheet_name': sheet_name,
                                'cell_reference': cell_ref,
                                'cell_value': cell_value,
                                'cell_formula': cell_formula,
                                'api_response': api_response
                            })
                            
                            # Small delay to avoid rate limiting
                            time.sleep(0.1)
                            
                        except Exception as e:
                            print(f"Error processing cell {cell_ref}: {e}")
                            results.append({
                                'sheet_name': sheet_name,
                                'cell_reference': cell_ref,
                                'cell_value': cell_value,
                                'cell_formula': cell_formula,
                                'api_response': f"Error: {str(e)}"
                            })
            
            return results
            
        except Exception as e:
            print(f"Error processing cell range: {e}")
            return []
    
    def save_results_to_excel(self, results: List[Dict], output_path: str):
        """
        Save results to a new Excel file
        
        Args:
            results: List of result dictionaries
            output_path: Path for the output Excel file
        """
        try:
            # Create DataFrame
            df = pd.DataFrame(results)
            
            # Reorder columns for better readability
            column_order = ['sheet_name', 'cell_reference', 'cell_value', 'cell_formula', 'api_response']
            df = df[column_order]
            
            # Save to Excel
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='AI_Processing_Results', index=False)
                
                # Auto-adjust column widths
                worksheet = writer.sheets['AI_Processing_Results']
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            print(f"Results saved to: {output_path}")
            
        except Exception as e:
            print(f"Error saving results: {e}")

def main():
    """Main function to run the processor"""
    print("=== Simple Excel AI Processor ===\n")
    
    # Get user inputs
    workbook_path = input("Enter the path to your Excel file: ").strip()
    
    if not os.path.exists(workbook_path):
        print("Error: File not found!")
        return
    
    # Load workbook to get sheet names
    try:
        wb = load_workbook(workbook_path)
        sheet_names = wb.sheetnames
        print(f"\nAvailable sheets: {', '.join(sheet_names)}")
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return
    
    sheet_name = input(f"Enter sheet name (or press Enter for '{sheet_names[0]}'): ").strip()
    if not sheet_name:
        sheet_name = sheet_names[0]
    
    cell_range = input("Enter cell range (e.g., A1:B10): ").strip()
    custom_prompt = input("Enter your custom prompt: ").strip()
    
    if not custom_prompt:
        custom_prompt = "Please analyze and process the following Excel cell content:"
    
    # Initialize processor
    processor = SimpleExcelAIProcessor()
    
    print(f"\nProcessing cells in range {cell_range} from sheet '{sheet_name}'...")
    
    # Process the cells
    results = processor.process_cell_range(workbook_path, sheet_name, cell_range, custom_prompt)
    
    if results:
        # Generate output filename
        base_name = os.path.splitext(os.path.basename(workbook_path))[0]
        output_path = f"{base_name}_AI_Results.xlsx"
        
        # Save results
        processor.save_results_to_excel(results, output_path)
        
        print(f"\nProcessing complete! Processed {len(results)} cells.")
        print(f"Results saved to: {output_path}")
    else:
        print("No results to save.")

if __name__ == "__main__":
    main() 