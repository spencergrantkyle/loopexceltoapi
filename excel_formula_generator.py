#!/usr/bin/env python3
"""
Excel Formula Generator
Converts text instructions with static cell references into dynamic Excel formulas.

This script is designed to process Excel files where text instructions reference 
specific cells, and convert those instructions into Excel formulas that will 
dynamically adjust when rows/columns are inserted or deleted.
"""

import os
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openai import OpenAI
from typing import List, Dict, Tuple, Set
import time
import re

# Load environment variables from .env file if it exists
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass  # dotenv is optional

class ExcelFormulaGenerator:
    def __init__(self, api_key: str = None):
        """
        Initialize the formula generator with OpenAI API key
        
        Args:
            api_key: OpenAI API key (if None, will try to get from environment)
        """
        if api_key:
            self.client = OpenAI(api_key=api_key)
        else:
            self.client = OpenAI()  # Will use OPENAI_API_KEY environment variable
        
        self.results = []
    
    def extract_cell_references(self, text: str) -> Set[str]:
        """
        Extract cell references from text using regex patterns
        
        Args:
            text: Text containing cell references
            
        Returns:
            Set of cell references found in the text
        """
        # Pattern to match Excel cell references (e.g., A1, AB123, $A$1, etc.)
        pattern = r'\b(?:\$?[A-Z]+\$?\d+)\b'
        matches = re.findall(pattern, text, re.IGNORECASE)
        
        # Clean up and standardize the matches
        clean_matches = set()
        for match in matches:
            # Remove $ signs for processing
            clean_ref = match.replace('$', '')
            clean_matches.add(clean_ref)
        
        return clean_matches
    
    def generate_dynamic_formula(self, text_instruction: str, cell_refs: Set[str], model: str = "gpt-4o") -> str:
        """
        Generate a dynamic Excel formula based on text instruction and cell references
        
        Args:
            text_instruction: Original text instruction
            cell_refs: Set of cell references found in the instruction
            model: OpenAI model to use
            
        Returns:
            Dynamic Excel formula as a string
        """
        try:
            # Create a prompt for the AI to convert static references to dynamic formulas
            cell_list = ", ".join(sorted(cell_refs))
            
            prompt = f"""
Convert the following text instruction into a dynamic Excel formula that will automatically adjust when rows or columns are inserted or deleted.

Original instruction: "{text_instruction}"

Cell references found: {cell_list}

Requirements:
1. Create a formula that expresses the same logic as the text instruction
2. Use Excel functions like CONCATENATE, INDIRECT, OFFSET, or similar to make references dynamic
3. The formula should work even if rows/columns are inserted or deleted
4. Return ONLY the Excel formula, starting with =
5. If the instruction involves validation or conditional logic, use appropriate Excel functions
6. If it's about cell formatting or data validation, create a formula that describes the validation rule

Example approaches:
- For text concatenation: =CONCATENATE("Text about cells: ", INDIRECT("F10"), ", ", INDIRECT("G10"))
- For validation rules: =IF(ISNUMBER(INDIRECT("F10")), "Valid", "Invalid")
- For range descriptions: ="Cells " & ADDRESS(ROW(F10),COLUMN(F10)) & " through " & ADDRESS(ROW(H20),COLUMN(H20)) & " should contain numeric values"

Please provide a formula that captures the essence of the instruction while being dynamic.
"""

            response = self.client.chat.completions.create(
                model=model,
                messages=[
                    {
                        "role": "system",
                        "content": "You are an Excel formula expert. You convert text instructions into dynamic Excel formulas that automatically adjust when spreadsheet structure changes. Always return only the Excel formula starting with =."
                    },
                    {
                        "role": "user",
                        "content": prompt
                    }
                ],
                max_tokens=500,
                temperature=0.2
            )
            
            formula = response.choices[0].message.content.strip()
            
            # Ensure the formula starts with =
            if not formula.startswith('='):
                formula = '=' + formula
                
            return formula
            
        except Exception as e:
            print(f"Error generating formula: {e}")
            return f"=ERROR(\"Failed to generate formula: {str(e)}\")"
    
    def process_instruction_range(self, workbook_path: str, sheet_name: str, 
                                instruction_column: str, instruction_range: str,
                                model: str = "gpt-4o") -> List[Dict]:
        """
        Process a range of cells containing text instructions and generate dynamic formulas
        
        Args:
            workbook_path: Path to the Excel file
            sheet_name: Name of the sheet
            instruction_column: Column containing text instructions (e.g., 'C')
            instruction_range: Range of rows to process (e.g., '1:10')
            model: OpenAI model to use
            
        Returns:
            List of results with original instructions and generated formulas
        """
        try:
            wb = load_workbook(workbook_path)
            ws = wb[sheet_name]
            
            # Parse the range
            if ':' in instruction_range:
                start_row, end_row = map(int, instruction_range.split(':'))
            else:
                start_row = end_row = int(instruction_range)
            
            results = []
            
            for row_num in range(start_row, end_row + 1):
                cell_ref = f"{instruction_column}{row_num}"
                cell = ws[cell_ref]
                instruction_text = str(cell.value) if cell.value else ""
                
                if instruction_text.strip():  # Only process non-empty cells
                    print(f"Processing instruction in cell {cell_ref}...")
                    
                    # Extract cell references from the instruction
                    cell_refs = self.extract_cell_references(instruction_text)
                    
                    if cell_refs:
                        # Generate dynamic formula
                        dynamic_formula = self.generate_dynamic_formula(
                            instruction_text, cell_refs, model
                        )
                        
                        results.append({
                            'sheet_name': sheet_name,
                            'instruction_cell': cell_ref,
                            'original_instruction': instruction_text,
                            'extracted_cell_refs': ', '.join(sorted(cell_refs)),
                            'dynamic_formula': dynamic_formula
                        })
                        
                        # Small delay to avoid rate limiting
                        time.sleep(0.1)
                    else:
                        results.append({
                            'sheet_name': sheet_name,
                            'instruction_cell': cell_ref,
                            'original_instruction': instruction_text,
                            'extracted_cell_refs': 'No cell references found',
                            'dynamic_formula': '=CONCATENATE("Text: ", "' + instruction_text.replace('"', '""') + '")'
                        })
            
            return results
            
        except Exception as e:
            print(f"Error processing instruction range: {e}")
            return []
    
    def save_formulas_to_excel(self, results: List[Dict], output_path: str):
        """
        Save generated formulas to a new Excel file
        
        Args:
            results: List of result dictionaries
            output_path: Path for the output Excel file
        """
        try:
            # Create DataFrame
            df = pd.DataFrame(results)
            
            # Reorder columns for better readability
            column_order = ['sheet_name', 'instruction_cell', 'original_instruction', 
                          'extracted_cell_refs', 'dynamic_formula']
            df = df[column_order]
            
            # Save to Excel
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Generated_Formulas', index=False)
                
                # Auto-adjust column widths
                worksheet = writer.sheets['Generated_Formulas']
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 80)  # Cap at 80 characters
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            print(f"Generated formulas saved to: {output_path}")
            
        except Exception as e:
            print(f"Error saving formulas: {e}")
    
    def copy_formulas_to_workbook(self, results: List[Dict], workbook_path: str, 
                                 target_sheet: str, target_column: str, 
                                 output_path: str = None):
        """
        Copy generated formulas directly to a specific column in the workbook
        
        Args:
            results: List of result dictionaries with formulas
            workbook_path: Path to the original Excel file
            target_sheet: Sheet name to write formulas to
            target_column: Column to write formulas to (e.g., 'D')
            output_path: Optional output path (if None, overwrites original)
        """
        try:
            wb = load_workbook(workbook_path)
            ws = wb[target_sheet]
            
            for result in results:
                # Extract row number from instruction cell (e.g., 'C1' -> 1)
                instruction_cell = result['instruction_cell']
                row_num = int(re.search(r'\d+', instruction_cell).group())
                
                # Write formula to target column
                target_cell = f"{target_column}{row_num}"
                ws[target_cell] = result['dynamic_formula']
                
                print(f"Added formula to {target_cell}")
            
            # Save the workbook
            save_path = output_path if output_path else workbook_path
            wb.save(save_path)
            print(f"Formulas added to workbook: {save_path}")
            
        except Exception as e:
            print(f"Error adding formulas to workbook: {e}")

def main():
    """Main function to run the formula generator"""
    print("=== Excel Formula Generator ===\n")
    print("This tool converts text instructions with cell references into dynamic Excel formulas.\n")
    
    # Get user inputs
    workbook_path = input("Enter the path to your Excel file: ").strip()
    
    if not os.path.exists(workbook_path):
        print("Error: File not found!")
        return
    
    # Load workbook to get sheet names
    try:
        wb = load_workbook(workbook_path)
        sheet_names = wb.sheetnames
        print(f"\nAvailable sheets: {', '.join(sheet_names)}")
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return
    
    sheet_name = input(f"Enter sheet name (or press Enter for '{sheet_names[0]}'): ").strip()
    if not sheet_name:
        sheet_name = sheet_names[0]
    
    instruction_column = input("Enter the column containing text instructions (e.g., C): ").strip().upper()
    instruction_range = input("Enter the row range to process (e.g., 1:10): ").strip()
    
    # Initialize generator
    generator = ExcelFormulaGenerator()
    
    print(f"\nProcessing instructions in column {instruction_column}, rows {instruction_range} from sheet '{sheet_name}'...")
    
    # Process the instructions
    results = generator.process_instruction_range(
        workbook_path, sheet_name, instruction_column, instruction_range
    )
    
    if results:
        # Ask user what they want to do with the results
        print(f"\nGenerated {len(results)} dynamic formulas.")
        print("\nWhat would you like to do?")
        print("1. Save formulas to a new Excel file for review")
        print("2. Add formulas directly to a column in your workbook")
        print("3. Both")
        
        choice = input("\nEnter your choice (1-3): ").strip()
        
        if choice in ['1', '3']:
            # Save to new file
            base_name = os.path.splitext(os.path.basename(workbook_path))[0]
            output_path = f"{base_name}_Generated_Formulas.xlsx"
            generator.save_formulas_to_excel(results, output_path)
        
        if choice in ['2', '3']:
            # Add to workbook
            target_column = input("Enter the column to add formulas to (e.g., D): ").strip().upper()
            create_copy = input("Create a copy of the workbook? (y/n): ").strip().lower()
            
            if create_copy == 'y':
                base_name = os.path.splitext(workbook_path)[0]
                output_path = f"{base_name}_with_formulas.xlsx"
            else:
                output_path = None
            
            generator.copy_formulas_to_workbook(
                results, workbook_path, sheet_name, target_column, output_path
            )
        
        print("\nProcessing complete!")
    else:
        print("No instructions found to process.")

if __name__ == "__main__":
    main()