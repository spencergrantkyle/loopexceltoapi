#!/usr/bin/env python3
"""
Example usage of Excel Formula Generator
Demonstrates how to convert text instructions with cell references into dynamic Excel formulas
"""

from excel_formula_generator import ExcelFormulaGenerator
import os

def example_data_validation_formulas():
    """Example: Convert data validation instructions to dynamic formulas"""
    
    # Initialize the generator
    generator = ExcelFormulaGenerator()
    
    # Example parameters based on the user's Excel structure
    workbook_path = "F.600.008 MFMA skills capacity and capability_Reimagined_DEV 1.xlsx"  # Replace with your file path
    sheet_name = "Sheet"  # Or whatever sheet contains your instructions
    instruction_column = "C"  # Column containing text instructions
    instruction_range = "1:10"  # Adjust based on your data range
    
    print("Processing data validation instructions...")
    
    # Process the instructions
    results = generator.process_instruction_range(
        workbook_path, sheet_name, instruction_column, instruction_range
    )
    
    # Save results to review
    if results:
        output_path = "data_validation_formulas.xlsx"
        generator.save_formulas_to_excel(results, output_path)
        print(f"Data validation formulas saved to: {output_path}")
        
        # Also copy to column D in the original workbook
        generator.copy_formulas_to_workbook(
            results, workbook_path, sheet_name, "D", 
            "workbook_with_dynamic_formulas.xlsx"
        )

def example_percentage_calculation_formulas():
    """Example: Convert percentage calculation instructions to dynamic formulas"""
    
    generator = ExcelFormulaGenerator()
    
    # For instructions like "Automate the calculation by calculating the percentage:"
    workbook_path = "your_workbook.xlsx"
    sheet_name = "Sheet"
    instruction_column = "C"
    instruction_range = "2:2"  # Just the second row with percentage instructions
    
    print("Processing percentage calculation instructions...")
    
    results = generator.process_instruction_range(
        workbook_path, sheet_name, instruction_column, instruction_range
    )
    
    if results:
        output_path = "percentage_formulas.xlsx"
        generator.save_formulas_to_excel(results, output_path)
        print(f"Percentage calculation formulas saved to: {output_path}")

def test_cell_reference_extraction():
    """Test the cell reference extraction functionality"""
    
    generator = ExcelFormulaGenerator()
    
    # Test with sample text from the user's Excel
    test_texts = [
        "The following cells should only allow numeric inputs: F10; G10; H10; F11; G11; H11; L11; F19; G19; H19; F20; G20 H20; L20; F25; G25; H25;F26; G26; H26; L26; F31; G31; H31; F32; G32; H32; L32; F37; G37; H37;F38; G38; H38;L38; F43; G43; H43; F44; G44; H44;L44",
        "Data validation Automate the calculation by calculating the percentage: - For F12 = F11/F10 - For F21 = F20/F19 - For F27 = F26F25 - For F33 = F32/F31 - For F39 = F38/F37 - For F44 = F43/F42 - For G12 = G11/G10 - For H12 = H11/H10 - For G21 = G20/G19 - For H21 = H20/H19 - For G27 = G26/G25"
    ]
    
    print("Testing cell reference extraction:\n")
    
    for i, text in enumerate(test_texts, 1):
        print(f"Test {i}:")
        print(f"Text: {text[:100]}...")
        cell_refs = generator.extract_cell_references(text)
        print(f"Extracted cell references: {sorted(cell_refs)}")
        print()

def create_sample_workbook():
    """Create a sample workbook with the structure shown in the user's Excel"""
    
    import openpyxl
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet"
    
    # Add headers and sample data
    ws['A1'] = "Sheet"
    ws['B1'] = "Cell Reference"
    ws['C1'] = "Text"
    ws['D1'] = "Formula with dynamic references"
    
    # Add sample data rows
    ws['A2'] = "Assessment_of_Org_Structure"
    ws['B2'] = "$Z$4"
    ws['C2'] = "The following cells should only allow numeric inputs: F10; G10; H10; F11; G11; H11; L11;"
    
    ws['A3'] = "Assessment_of_Org_Structure"
    ws['B3'] = "$Z$5"
    ws['C3'] = "Data validation Automate the calculation by calculating the percentage: - For F12 = F11/F10 - For F21 = F20/F19"
    
    # Save the sample workbook
    wb.save("sample_workbook.xlsx")
    print("Sample workbook created: sample_workbook.xlsx")

if __name__ == "__main__":
    print("=== Excel Formula Generator Examples ===\n")
    
    # Check if OpenAI API key is set
    if not os.getenv("OPENAI_API_KEY"):
        print("Warning: OPENAI_API_KEY environment variable not set!")
        print("Please set your OpenAI API key before running examples.\n")
    
    print("Available examples:")
    print("1. Test cell reference extraction")
    print("2. Create sample workbook")
    print("3. Process data validation instructions (requires your Excel file)")
    print("4. Process percentage calculation instructions (requires your Excel file)")
    
    choice = input("\nEnter your choice (1-4) or press Enter to exit: ").strip()
    
    if choice == "1":
        test_cell_reference_extraction()
    elif choice == "2":
        create_sample_workbook()
    elif choice == "3":
        example_data_validation_formulas()
    elif choice == "4":
        example_percentage_calculation_formulas()
    else:
        print("Exiting...")