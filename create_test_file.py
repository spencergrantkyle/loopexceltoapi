#!/usr/bin/env python3
"""
Create a test Excel file for demonstrating the Simple Excel AI Processor
"""

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def create_test_excel():
    """Create a test Excel file with sample data"""
    
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "TestData"
    
    # Sample data
    test_data = [
        ["Hello World", "=A1", "Welcome to our company"],
        ["Goodbye", "=A2", "Thank you for your business"],
        ["Financial Report", "=A3", "Q4 2024 Results"],
        ["Revenue", "=SUM(B1:B3)", "$1,250,000"],
        ["Expenses", "=SUM(B4:B6)", "$850,000"],
        ["Profit", "=B7-B8", "$400,000"],
        ["Growth Rate", "=(B9/B7)*100", "32%"],
        ["Customer Satisfaction", "=AVERAGE(B10:B12)", "4.5/5"],
        ["Employee Count", "=COUNT(B13:B15)", "150"],
        ["Project Status", "=IF(B16>100,\"On Track\",\"Behind\")", "On Track"]
    ]
    
    # Add headers
    headers = ["Text Content", "Formula", "Description"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Add data
    for row_idx, row_data in enumerate(test_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # If it's a formula (starts with =), set it as formula
            if isinstance(value, str) and value.startswith('='):
                cell.value = value
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the file
    filename = "test_data.xlsx"
    wb.save(filename)
    print(f"Test Excel file created: {filename}")
    print(f"Contains {len(test_data)} rows of sample data with text, formulas, and descriptions")
    
    return filename

if __name__ == "__main__":
    create_test_excel() 